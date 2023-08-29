import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import points_to_pixels, pixels_to_EMU

from lrb.common import log

CHARACTER_WIDTH = 6


def insert_image(sheet, cell, image_coordinate):
    column_width = 30
    row_height = 195
    sheet.column_dimensions[cell.column_letter].width = column_width
    sheet.row_dimensions[cell.row].height = row_height
    column_pixels = points_to_pixels(column_width * CHARACTER_WIDTH) - 10
    row_pixels = points_to_pixels(row_height) - 10

    image_name = str(sheet[image_coordinate].value) + '.jpg'
    log.info('插入图片：[{}]，坐标：[{}]，列宽：[{}]，行高：[{}]', image_name, cell.coordinate, column_pixels, row_pixels)
    image_path = f'C:\\Users\\Magese\\Desktop\\addimg\\imgs\\{image_name}'
    img = Image(image_path)

    img_height = img.height
    img_width = img.width

    resize_factor = min(column_pixels / img_width, row_pixels / img_height)
    resize_w = img_width * resize_factor
    resize_h = img_height * resize_factor

    img.height = resize_h
    img.width = resize_w

    x = 5
    y = 20

    img_size_excel = XDRPositiveSize2D(pixels_to_EMU(resize_w), pixels_to_EMU(resize_h))
    marker = AnchorMarker(col=cell.column - 1, colOff=pixels_to_EMU(x), row=cell.row - 1, rowOff=pixels_to_EMU(y))
    img.anchor = OneCellAnchor(_from=marker, ext=img_size_excel)

    sheet.add_image(img)


def main():
    excel_path = r'C:\Users\Magese\Desktop\addimg\【每日更新】LRB SEM post －test.xlsx'
    xlsx = openpyxl.load_workbook(excel_path)
    active = xlsx.active
    max_row = active.max_row
    max_column = active.max_column

    for i in range(1, max_row + 1):
        for j in range(1, max_column + 1):
            cell = active.cell(row=i, column=j)
            if str(cell.value).startswith('="<table>'):
                value = str(cell.value)
                value = value[(value.index('"&') + 2):value.index('&"')]
                insert_image(active, cell, value)
    xlsx.save(excel_path)


if __name__ == '__main__':
    main()
