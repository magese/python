import openpyxl


class Item:
    row = 0
    id = ''
    link = ''
    log = ''

    def __init__(self, r, i, l):
        self.row = r
        self.id = i
        self.link = l

    def to_string(self):
        return 'row=%s, id=%s, link=%s' % (self.row, self.id, self.link)


filepath = 'C:\\Users\\Magese\\Desktop\\lrb_demo.xlsx'
xlsx = openpyxl.load_workbook(filepath)

sheet = xlsx.active
max_row = sheet.max_row
max_column = sheet.max_column
print('最大行数：', max_row, '，最大列数：', max_column)

items = []
for i in range(1, max_row + 1):
    id = sheet.cell(row=i, column=1).value
    link = sheet.cell(row=i, column=2).value
    item = Item(i, id, link)
    items.append(item)
    print(item.to_string())

for i in range(0, len(items)):
    item = items[i]
    sheet.cell(row=item.row, column=3, value='success')

xlsx.save(filepath)
print('finish')