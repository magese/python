import base64
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.wait import WebDriverWait


class Item:
    row = 0
    link = ''
    name = ''

    def __init__(self, _row, _link, _name):
        self.row = _row
        self.link = _link
        self.name = _name

    def to_string(self):
        return 'row=%s, link=%s, name=%s' % (self.row, self.link, self.name)


# 打开浏览器
def open_browser():
    options = Options()
    options.add_experimental_option('detach', True)
    options.add_argument(r'--user-data-dir=C:\Users\Magese\AppData\Local\Microsoft\Edge\User Data')
    return webdriver.Edge(options=options)


# 读取excel
def read_excel(sheet):
    max_row = sheet.max_row
    max_column = sheet.max_column
    print('读取excel，最大行数：', max_row, '，最大列数：', max_column)

    items = []
    for j in range(1, max_row + 1):
        if str(sheet.cell(row=j, column=3).value).startswith('success'):
            continue
        item = Item(j, sheet.cell(row=j, column=1).value, sheet.cell(row=j, column=2).value)
        items.append(item)
    return items


# 等待查找元素
def wait_for_find_ele(edge, func):
    ele = WebDriverWait(edge, timeout=10).until(func)
    time.sleep(0.5)
    return ele


# base64转图片
def base64_to_image_file(base64_image, file_path):
    img_data = base64.b64decode(base64_image)
    with open(file_path, 'wb') as f:
        f.write(img_data)


# 截图
def screenshot(edge, row, save_dir):
    edge.get(row.link)
    edge.maximize_window()
    time.sleep(2)
    filename = row.name + '.jpg'
    jpg_path = save_dir + filename
    note = wait_for_find_ele(edge, lambda d: d.find_element(by=By.ID, value='noteContainer'))
    base64_img = note.screenshot_as_base64
    base64_to_image_file(base64_img, jpg_path)
    print('save jpg success', jpg_path)


# main
filepath = 'C:\\Users\\Magese\\Desktop\\LRB笔记截图.xlsx'
xlsx = openpyxl.load_workbook(filepath)
active = xlsx.active
lines = read_excel(active)
size = len(lines)
print('共读取待截图记录', size, '条')

drive = open_browser()
jpg_dir = 'C:\\Users\\Magese\\Desktop\\lrb_screenshot\\'

for i in range(0, size):
    line = lines[i]
    screenshot(drive, line, jpg_dir)
drive.close()
