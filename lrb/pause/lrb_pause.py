import time

import openpyxl
from selenium.common import StaleElementReferenceException
from selenium.webdriver.common.by import By

from lrb.common import log, util
from lrb.common.Model import Excel, Lrb


class Item:
    row = 0
    id = ''
    log = ''

    def __init__(self, _row, _id):
        self.row = _row
        self.id = _id

    def to_string(self):
        return 'row=%s, id=%s' % (self.row, self.id)


class LrbPause(Lrb):

    def __read_excel(self):
        xlsx = openpyxl.load_workbook(self._excel_path)
        active = xlsx.active
        max_row = active.max_row
        max_column = active.max_column
        self.msg.emit(log.msg('读取excel最大行数：{}，最大列数：{}', max_row, max_column))

        items = []
        for j in range(1, max_row + 1):
            if str(active.cell(row=j, column=2).value).startswith('success'):
                continue
            item = Item(j, active.cell(row=j, column=1).value)
            items.append(item)
        self.msg.emit(log.msg('共读取待暂停笔记ID {}条', len(items)))
        self.excel = Excel(self._excel_path, xlsx, active, max_row, max_column, items)

    def __pause_note(self):
        retry = 3
        is_paused = True
        while retry > 0:
            try:
                util.search_id(self.item.id, self.edge)
                time.sleep(1)

                pause_checkbox = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="css-1a4w089"), self.edge)
                checked = pause_checkbox.get_attribute('data-is-checked')

                pause_switch = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="css-vwjppn"), self.edge)

                if 'true' == checked:
                    pause_switch.click()
                    is_paused = False
                time.sleep(0.5)
                break
            except StaleElementReferenceException:
                retry -= 1
                time.sleep(0.5)
        return is_paused

    def run(self):
        try:
            super().execute(
                action='暂停笔记',
                read_func=self.__read_excel,
                page_func=util.creative_page,
                do_func=self.__pause_note,
                end_func=None,
                res_column=2
            )
        except BaseException as e:
            self._emit('发生未知异常，错误信息：{}', repr(e))
            self._err('发生未知异常，错误信息：{}', repr(e))


# main
# noinspection PyUnresolvedReferences
def main():
    lp = LrbPause()
    lp._excel_path = r'C:\Users\mages\Desktop\暂停创意id.xlsx'
    lp._username = ''
    lp._password = ''
    lp.msg.connect(lambda m: print(m))
    lp.run()


if __name__ == '__main__':
    main()
