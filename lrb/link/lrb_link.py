import time

import openpyxl
from selenium.common import StaleElementReferenceException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By

from lrb.common import util
from lrb.common.Model import Lrb, Excel


class Item:
    row = 0
    id = ''
    link = ''
    log = ''

    def __init__(self, _row, _id, _link):
        self.row = _row
        self.id = _id
        self.link = _link

    def to_string(self):
        return 'row=%s, id=%s' % (self.row, self.id)


class LrbLink(Lrb):

    def __read_excel(self):
        xlsx = openpyxl.load_workbook(self._excel_path)
        active = xlsx.active
        max_row = active.max_row
        max_column = active.max_column
        self._emit('读取excel最大行数：{}，最大列数：{}', max_row, max_column)

        items = []
        for j in range(1, max_row + 1):
            if str(active.cell(row=j, column=3).value).startswith('success'):
                continue
            item = Item(j, active.cell(row=j, column=1).value, active.cell(row=j, column=2).value)
            items.append(item)
        self._emit('共读取待更换监测链接记录 {}条', len(items))
        self.excel = Excel(self._excel_path, xlsx, active, max_row, max_column, items)

    def __change_link(self):
        util.search_id(self.item.id, self.edge)
        time.sleep(1.2)

        edit_retry = 3
        while edit_retry > 0:
            try:
                util.click_edit(self.edge)
                time.sleep(1.2)
                break
            except StaleElementReferenceException:
                edit_retry -= 1
                time.sleep(0.5)

        self.edge.switch_to.window(self.edge.window_handles[-1])

        table = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="el-table__fixed-body-wrapper"), self.edge)
        td = util.wait_for_find_ele(
            lambda d: table.find_elements(by=By.TAG_NAME, value="td"), self.edge)
        change_btn = util.wait_for_find_ele(
            lambda d: td[13].find_element(by=By.CLASS_NAME, value="link-text"), self.edge)
        change_btn.click()

        clear_btn = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-1jjt3ne"), self.edge)
        clear_btn.click()
        link_input = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-968ze5"), self.edge)
        link_input.clear()
        link_input.send_keys('')
        link_input.send_keys(Keys.CONTROL, 'a')
        link_input.send_keys(self.item.link)

        save_btn = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-php29w"), self.edge)
        save_btn.click()
        time.sleep(0.7)

        finish_btn = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-r7neow"), self.edge)
        finish_btn.click()
        time.sleep(1)
        self.edge.close()
        self.edge.switch_to.window(self.edge.window_handles[0])

    def run(self):
        try:
            super().execute(
                action='监测链接修改',
                read_func=self.__read_excel,
                page_func=util.creative_page,
                do_func=self.__change_link,
                end_func=None,
                res_column=3
            )
        except BaseException as e:
            self._emit('发生未知异常，错误信息：{}', repr(e))
            self._err('发生未知异常，错误信息：{}', repr(e))


def main():
    ll = LrbLink()
    ll._excel_path = r'C:\Users\mages\Desktop\创意id+监测链接.xlsx'
    ll._username = ''
    ll._password = ''
    # noinspection PyUnresolvedReferences
    ll.msg.connect(lambda m: print(m))
    ll.run()


if __name__ == '__main__':
    main()
