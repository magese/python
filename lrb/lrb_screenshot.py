import base64
import time
import traceback

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options

import lrb_util


class Item:
    row = 0
    link = ''
    name = ''

    def __init__(self, _row, _link, _name):
        self.row = _row
        self.link = _link
        self.name = _name

    def to_string(self):
        return 'row=%s, link=%s, name=%s' % (self.row, self.link, self.name)


# 打开浏览器
def open_browser():
    options = Options()
    options.add_experimental_option('detach', True)
    options.add_argument(r'--user-data-dir=C:\Users\Magese\AppData\Local\Microsoft\Edge\User Data')
    return webdriver.Edge(options=options)


# 读取excel
def read_excel(sheet):
    max_row = sheet.max_row
    max_column = sheet.max_column
    lrb_util.log('读取excel，最大行数：{}，最大列数：{}', max_row, max_column)

    items = []
    for j in range(1, max_row + 1):
        if str(sheet.cell(row=j, column=3).value).startswith('success'):
            continue
        item = Item(j, sheet.cell(row=j, column=1).value, sheet.cell(row=j, column=2).value)
        items.append(item)
    return items


# base64转图片
def base64_to_image_file(base64_image, file_path):
    img_data = base64.b64decode(base64_image)
    with open(file_path, 'wb') as f:
        f.write(img_data)


# 截图
def screenshot(edge, row, save_dir):
    edge.get(row.link)
    edge.maximize_window()
    time.sleep(2)
    filename = row.name + '.jpg'
    jpg_path = save_dir + filename
    note = lrb_util.wait_for_find_ele(edge, lambda d: d.find_element(by=By.ID, value='noteContainer'))
    base64_img = note.screenshot_as_base64
    base64_to_image_file(base64_img, jpg_path)
    return jpg_path


# main
filepath = 'C:\\Users\\Magese\\Desktop\\LRB笔记截图.xlsx'
xlsx = openpyxl.load_workbook(filepath)
active = xlsx.active
lines = read_excel(active)
size = len(lines)
lrb_util.log('共读取待截图记录{}条', size)

drive = open_browser()
jpg_dir = 'C:\\Users\\Magese\\Desktop\\lrb_screenshot\\'
success = 0
failure = 0
result = ''
start = time.perf_counter()

for i in range(0, size):
    line = lines[i]
    try:
        save_path = screenshot(drive, line, jpg_dir)
        lrb_util.log('{} => save jpg success => {}', lrb_util.loop_msg(i + 1, size, start), save_path)
    except BaseException as e:
        traceback.print_exc()
        failure += 1
        result = 'failure:' + repr(e)
        lrb_util.log('{} => save jpg failure => {}', lrb_util.loop_msg(i + 1, size, start), result)
    else:
        success += 1
        result = 'success'
    finally:
        start = time.perf_counter()
        active.cell(row=line.row, column=3, value=result)
        xlsx.save(filepath)

lrb_util.log('截图任务完成，成功{}条，失败{}条', success, failure)
drive.close()