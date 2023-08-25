import time

import openpyxl
from selenium.common import StaleElementReferenceException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By

from lrb.common import util
from lrb.common.Model import Excel, Lrb


class Item:
    row = 0
    id = ''
    name = ''
    log = ''

    def __init__(self, _row, _id, _name):
        self.row = _row
        self.id = _id
        self.name = _name

    def to_string(self):
        return 'row=%s, id=%s, name=%s' % (self.row, self.id, self.name)


class LrbNoteName(Lrb):

    def __read_excel(self):
        xlsx = openpyxl.load_workbook(self._excel_path)
        active = xlsx.active
        max_row = active.max_row
        max_column = active.max_column
        self._emit('读取excel最大行数：{}，最大列数：{}', max_row, max_column)

        items = []
        for j in range(1, max_row + 1):
            if str(active.cell(row=j, column=3).value).startswith('success'):
                continue
            item = Item(j, active.cell(row=j, column=1).value, active.cell(row=j, column=2).value)
            items.append(item)
        self._emit('共读取待修改名称记录 {}条', len(items))
        self.excel = Excel(self._excel_path, xlsx, active, max_row, max_column, items)

    def __change_note_name(self):
        util.search_id(self.item.id, self.edge)
        time.sleep(1.2)

        tbody = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="d-table__body"), self.edge)
        tr = util.wait_for_find_ele(
            lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)
        td = util.wait_for_find_ele(
            lambda d: tr[0].find_elements(by=By.TAG_NAME, value="td"), self.edge)
        span = util.wait_for_find_ele(
            lambda d: td[3].find_element(by=By.TAG_NAME, value="span"), self.edge)
        creative_name = span.get_attribute('innerText')
        if self.item.name == creative_name:
            time.sleep(0.5)
            return True

        edit_retry = 3
        while edit_retry > 0:
            try:
                util.click_edit(self.edge)
                time.sleep(1)
                break
            except StaleElementReferenceException:
                edit_retry -= 1
                time.sleep(0.5)

        self.edge.switch_to.window(self.edge.window_handles[-1])

        name_input = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-1azanbt"), self.edge)
        input_value = name_input.get_attribute('value')
        if input_value == self.item.name:
            self.edge.close()
            self.edge.switch_to.window(self.edge.window_handles[0])
            return True

        clear_btn = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-18cbzsm"), self.edge)
        clear_btn.click()

        name_input.send_keys('')
        name_input.clear()
        name_input.send_keys(Keys.CONTROL, 'a')
        name_input.send_keys(self.item.name)
        time.sleep(0.5)

        finish_btn = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-r7neow"), self.edge)
        finish_btn.click()
        time.sleep(1)
        self.edge.close()
        self.edge.switch_to.window(self.edge.window_handles[0])
        return False

    def run(self):
        try:
            super().execute(
                '名称修改',
                self.__read_excel,
                util.creative_page,
                self.__change_note_name,
                3
            )
        except BaseException as e:
            self._emit('发生未知异常，错误信息：{}', repr(e))
            self._err('发生未知异常，错误信息：{}', repr(e))


# main
def main():
    lnn = LrbNoteName()
    lnn._excel_path = r'C:\Users\Magese\Desktop\创意名称修改.xlsx'
    lnn._username = ''
    lnn._password = ''
    # noinspection PyUnresolvedReferences
    lnn.msg.connect(lambda m: print(m))
    lnn.run()


if __name__ == '__main__':
    main()
