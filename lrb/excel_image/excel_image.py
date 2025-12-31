import os
import traceback

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import points_to_pixels, pixels_to_EMU

from lrb.common.Model import Excel, Lrb


class Item:
    sheet = None
    cell = None
    img_name: str = ''
    coordinate: str = ''

    def __init__(self, _sheet, _cell, _img_name, _coordinate):
        self.sheet = _sheet
        self.cell = _cell
        self.img_name = _img_name
        self.coordinate = _coordinate

    def to_string(self):
        return 'coordinate=%s, img_name=%s' % (self.coordinate, self.img_name)


class ExcelImage(Lrb):
    CHARACTER_WIDTH = 6

    def __read_excel(self):
        xlsx = openpyxl.load_workbook(self._excel_path)
        sheets = xlsx.sheetnames
        items = []
        for sheet in sheets:
            active = xlsx[sheet]
            max_row = active.max_row
            max_column = active.max_column
            self._emit('读取excel，sheet名称：{}，最大行数：{}，最大列数：{}', sheet, max_row, max_column)

            for i in range(1, max_row + 1):
                for j in range(1, max_column + 1):
                    cell = active.cell(row=i, column=j)
                    if str(cell.value).startswith('="<table>'):
                        value = str(cell.value)
                        coordinate = value[(value.index('"&') + 2):value.index('&"')]
                        image_name = str(active[coordinate].value) + '.jpg'
                        items.append(Item(_sheet=active, _cell=cell, _img_name=image_name, _coordinate=coordinate))
            self._emit('共读取待插入图片单元格 {}个', len(items))
        self.excel = Excel(self._excel_path, xlsx, None, 0, 0, items)

    def __nothing(self, *p):
        pass

    def __add_image(self):
        column_width = 25
        row_height = 180
        self.item.sheet.column_dimensions[self.item.cell.column_letter].width = column_width
        self.item.sheet.row_dimensions[self.item.cell.row].height = row_height
        column_pixels = points_to_pixels(column_width * self.CHARACTER_WIDTH) - 10
        row_pixels = points_to_pixels(row_height) - 10

        image_path = f'{os.path.abspath(os.path.join(self._excel_path, os.pardir))}\\{self.item.img_name}'
        img = Image(image_path)

        img_height = img.height
        img_width = img.width

        resize_factor = min(column_pixels / img_width, row_pixels / img_height)
        resize_w = img_width * resize_factor
        resize_h = img_height * resize_factor

        img.height = resize_h
        img.width = resize_w

        x = 5
        y = 40

        img_size_excel = XDRPositiveSize2D(pixels_to_EMU(resize_w), pixels_to_EMU(resize_h))
        marker = AnchorMarker(col=self.item.cell.column - 1, colOff=pixels_to_EMU(x),
                              row=self.item.cell.row - 1, rowOff=pixels_to_EMU(y))
        img.anchor = OneCellAnchor(_from=marker, ext=img_size_excel)

        self.item.sheet.add_image(img)

    def _save_excel(self):
        self.excel.xlsx.save(self.excel.path)

    def run(self):
        try:
            super().execute(
                action='插入Excel图片',
                read_func=self.__read_excel,
                page_func=self.__nothing,
                do_func=self.__add_image,
                end_func=self._save_excel,
                res_column=-1,
                interval_ms=0
            )
        except BaseException as e:
            self._emit('发生未知异常，错误信息：{}', repr(e))
            self._err('发生未知异常，错误信息：{}', repr(e))
            traceback.print_exc()


def main():
    ei = ExcelImage()
    ei._excel_path = r'C:\Users\Magese\Desktop\screenshot2\冲锋衣小红书周报-1028-1104.xlsx'
    # noinspection PyUnresolvedReferences
    ei.msg.connect(lambda m: print(m))
    ei.run()


if __name__ == '__main__':
    main()
