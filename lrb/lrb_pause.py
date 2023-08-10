import time

import openpyxl
from selenium.common import StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from lrb.common import log, util
from lrb.common.Model import Excel, Lrb


class Item:
    row = 0
    id = ''
    log = ''

    def __init__(self, _row, _id):
        self.row = _row
        self.id = _id

    def to_string(self):
        return 'row=%s, id=%s' % (self.row, self.id)


class LrbPause(Lrb):

    def __read_excel(self):
        xlsx = openpyxl.load_workbook(self._excel_path)
        active = xlsx.active
        max_row = active.max_row
        max_column = active.max_column
        self.msg.emit(log.msg('读取excel最大行数：{}，最大列数：{}', max_row, max_column))

        items = []
        for j in range(1, max_row + 1):
            if str(active.cell(row=j, column=2).value).startswith('success'):
                continue
            item = Item(j, active.cell(row=j, column=1).value)
            items.append(item)
        self.msg.emit(log.msg('共读取待暂停笔记ID {}条', len(items)))
        self.excel = Excel(self._excel_path, xlsx, active, max_row, max_column, items)

    def __pause_note(self):
        retry = 3
        is_paused = True
        while retry > 0:
            try:
                manage = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="manage-list"), self.edge)
                id_input = util.wait_for_find_ele(
                    lambda d: manage.find_element(by=By.TAG_NAME, value="input"), self.edge)
                id_input.send_keys('')
                id_input.clear()
                id_input.send_keys(self.item.id)
                id_input.send_keys(Keys.ENTER)
                time.sleep(1)

                pause_checkbox = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="css-1a4w089"), self.edge)
                checked = pause_checkbox.get_attribute('data-is-checked')

                pause_switch = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="css-vwjppn"), self.edge)

                if 'true' == checked:
                    pause_switch.click()
                    is_paused = False
                time.sleep(0.5)
                break
            except StaleElementReferenceException:
                retry -= 1
                time.sleep(0.5)
        return is_paused

    def begin(self):
        super().execute(
            '暂停笔记',
            self.__read_excel,
            util.unit_page,
            self.__pause_note,
            2
        )


# main
# noinspection PyUnresolvedReferences
def main():
    username = ''
    password = ''
    filepath = 'C:\\Users\\mages\\Desktop\\暂停创意id.xlsx'
    lp = LrbPause(filepath, username, password)
    lp.msg.connect(lambda m: print(m))
    lp.begin()


if __name__ == '__main__':
    main()
