import time

import openpyxl
from selenium.common import StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By

from lrb.common import util
from lrb.common.Model import Excel, Lrb


class Item:
    row: int = 0
    id: str = ''
    exact_word: str = ''
    fuzzy_word: str = ''
    exact_len: int = 0
    fuzzy_len: int = 0

    def __init__(self, _row, _id, _exact_word, _fuzzy_word):
        self.row = _row
        self.id = _id
        self.exact_word = _exact_word
        self.fuzzy_word = _fuzzy_word
        self.exact_len = len(_exact_word.split('\n')) if _exact_word else 0
        self.fuzzy_len = len(_fuzzy_word.split('\n')) if _fuzzy_word else 0

    def to_string(self):
        return ('row=%s, id=%s, exact_word.size=%s, fuzzy_word.size=%s' %
                (self.row, self.id, self.exact_len, self.fuzzy_len))


class LrbNegativeWord(Lrb):

    def __read_excel(self):
        xlsx = openpyxl.load_workbook(self._excel_path)
        active = xlsx.active
        max_row = active.max_row
        max_column = active.max_column
        self._emit('读取excel最大行数：{}，最大列数：{}', max_row, max_column)

        items = []
        for j in range(1, max_row + 1):
            if str(active.cell(row=j, column=4).value).startswith('success'):
                continue
            item = Item(j, active.cell(row=j, column=1).value,
                        active.cell(row=j, column=2).value, active.cell(row=j, column=3).value)
            items.append(item)
        self._emit('共读取待添加否定词记录 {}条', len(items))
        self.excel = Excel(self._excel_path, xlsx, active, max_row, max_column, items)

    def __add_word(self):
        util.search_id(self.item.id, self.edge)
        time.sleep(1.2)

        retry = 3
        while retry > 0:
            try:
                tbody = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="d-table__body"), self.edge)
                tr = util.wait_for_find_ele(
                    lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)
                td = util.wait_for_find_ele(
                    lambda d: tr[0].find_elements(by=By.TAG_NAME, value="td"), self.edge)
                a = util.wait_for_find_ele(
                    lambda d: td[2].find_element(by=By.TAG_NAME, value="a"), self.edge)
                a.click()
                time.sleep(1.2)
                break
            except StaleElementReferenceException:
                retry -= 1
                time.sleep(0.5)

        retry = 3
        while retry > 0:
            try:
                tabs_headers = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="d-tabs-headers-linear"), self.edge)
                tabs = util.wait_for_find_ele(
                    lambda d: tabs_headers.find_elements(by=By.CLASS_NAME, value="d-tabs-header-linear"), self.edge)
                tabs[2].click()
                time.sleep(1)
                break
            except StaleElementReferenceException:
                retry -= 1
                time.sleep(0.5)

        self.__de_duplicate()
        if self.item.exact_len == 0 and self.item.fuzzy_len == 0:
            time.sleep(1)
            util.switch_unit_id_page(self.edge)
            return True

        if self.item.exact_len > 0:
            self.__do_add(self.item.exact_word, 0)

        if self.item.fuzzy_len > 0:
            self.__do_add(self.item.fuzzy_word, 1)

        util.switch_unit_id_page(self.edge)

        return False

    def __do_add(self, words, radio_idx):
        manage = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="manage-list"), self.edge)
        btn = util.wait_for_find_ele(
            lambda d: manage.find_elements(by=By.TAG_NAME, value="button"), self.edge)
        btn[0].click()
        time.sleep(0.5)

        modal = util.wait_for_find_ele(
            lambda d: d.find_elements(by=By.CLASS_NAME, value="d-modal"), self.edge)[0]
        radios = util.wait_for_find_ele(
            lambda d: modal.find_elements(by=By.CLASS_NAME, value="d-radio-simulator"), self.edge)
        textarea = util.wait_for_find_ele(
            lambda d: modal.find_element(by=By.TAG_NAME, value="textarea"), self.edge)
        btns = util.wait_for_find_ele(
            lambda d: modal.find_elements(by=By.TAG_NAME, value="button"), self.edge)

        radios[radio_idx].click()
        textarea.clear()
        textarea.send_keys(words)
        time.sleep(1)
        btns[1].click()
        time.sleep(1)

        try:
            cancel_btn = util.wait_for_find_ele(
                lambda d: d.find_element(by=By.CLASS_NAME, value="css-fm44j"), self.edge, 1)
            cancel_btn.click()
        except TimeoutException:
            pass

    def __de_duplicate(self):
        tbody = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="d-table__body"), self.edge)
        tr = util.wait_for_find_ele(
            lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)
        if len(tr) < 2:
            return

        # page_select = util.wait_for_find_ele(
        #     lambda d: d.find_element(by=By.CLASS_NAME, value="d-select-main-indicator"), self.edge)
        # page_select.click()
        # time.sleep(0.5)
        #
        # dropdown = util.wait_for_find_ele(
        #     lambda d: d.find_element(by=By.CLASS_NAME, value="d-dropdown-content"), self.edge)
        # grid_item = util.wait_for_find_ele(
        #     lambda d: dropdown.find_elements(by=By.CLASS_NAME, value="d-grid-item"), self.edge)
        # grid_item[len(grid_item) - 1].click()
        # time.sleep(1)
        tr = util.wait_for_find_ele(
            lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)

        exact_words = self.item.exact_word.split('\n') if self.item.exact_word else []
        fuzzy_words = self.item.fuzzy_word.split('\n') if self.item.fuzzy_word else []
        for i in range(1, len(tr)):
            td = util.wait_for_find_ele(
                lambda d: tr[i].find_elements(by=By.TAG_NAME, value="td"), self.edge)
            word = td[1].get_attribute('innerText')
            meth = td[2].get_attribute('innerText')
            if meth == '精确' and word in exact_words:
                exact_words.remove(word)
            elif meth == '短语' and word in fuzzy_words:
                fuzzy_words.remove(word)
        self.item.exact_word = '\n'.join(exact_words) if len(exact_words) > 0 else None
        self.item.fuzzy_word = '\n'.join(fuzzy_words) if len(fuzzy_words) > 0 else None
        self.item.exact_len = len(exact_words)
        self.item.fuzzy_len = len(fuzzy_words)

    def run(self):
        try:
            super().execute(
                action='添加否定词',
                read_func=self.__read_excel,
                page_func=util.unit_page,
                do_func=self.__add_word,
                end_func=None,
                res_column=4
            )
        except BaseException as e:
            self._emit('发生未知异常，错误信息：{}', repr(e))
            self._err('发生未知异常，错误信息：{}', repr(e))


def main():
    lnw = LrbNegativeWord()
    lnw._excel_path = r'C:\Users\mages\Desktop\批量否词需求-1215-r.xlsx'
    lnw._username = 'skiicn_lrb2021@163.com'
    lnw._password = 'Mediacom12345'
    # noinspection PyUnresolvedReferences
    lnw.msg.connect(lambda m: print(m))
    lnw.run()


if __name__ == '__main__':
    main()
