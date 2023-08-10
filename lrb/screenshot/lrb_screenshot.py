import base64
import time
import traceback

import openpyxl
from PyQt6.QtCore import QThread, pyqtSignal
from selenium.webdriver.common.by import By

from lrb.common import log, util
from lrb.common.Model import Excel


class Item:
    row = 0
    link = ''
    name = ''

    def __init__(self, _row, _link, _name):
        self.row = _row
        self.link = _link
        self.name = _name

    def to_string(self):
        return 'row=%s, link=%s, name=%s' % (self.row, self.link, self.name)


class LrbScreenshot(QThread):
    excel: Excel = None
    stop_status: int = 0
    excel_path: str = ''
    save_dir: str = ''

    status = pyqtSignal(int)
    value_changed = pyqtSignal(float)
    output = pyqtSignal(str)

    def __read_excel(self):
        try:
            xlsx = openpyxl.load_workbook(self.excel_path)
            active = xlsx.active
            max_row = active.max_row
            max_column = active.max_column
            self.__log('读取excel最大行数：{}，最大列数：{}', max_row, max_column)

            items = []
            for j in range(1, max_row + 1):
                if str(active.cell(row=j, column=3).value).startswith('success'):
                    continue
                item = Item(j, active.cell(row=j, column=1).value, active.cell(row=j, column=2).value)
                items.append(item)
            self.__log('共读取待截图记录{}条', len(items))
            self.excel = Excel(self.excel_path, xlsx, active, max_row, max_column, items)
        except BaseException as e:
            traceback.print_exc()
            self.__log('读取excel失败，请检查文件，错误信息：{}', repr(e))
            return

    @staticmethod
    def __base64_to_image_file(base64_image, file_path):
        img_data = base64.b64decode(base64_image)
        with open(file_path, 'wb') as f:
            f.write(img_data)

    def __screenshot(self, edge, row):
        edge.get(row.link)
        time.sleep(2)
        filename = row.name + '.jpg'
        jpg_path = self.save_dir + '/' + filename
        note = util.wait_for_find_ele(lambda d: d.find_element(by=By.ID, value='noteContainer'), edge)
        base64_img = note.screenshot_as_base64
        self.__base64_to_image_file(base64_img, jpg_path)
        return jpg_path

    # noinspection PyUnresolvedReferences
    def __log(self, k, *p):
        msg = log.msg(k, *p)
        self.output.emit(msg)

    # noinspection PyUnresolvedReferences
    def exec(self):
        self.__read_excel()

        size = len(self.excel.lines)
        if size <= 0:
            return

        try:
            drive = util.open_browser()
            drive.maximize_window()
        except BaseException as e:
            traceback.print_exc()
            self.__log('打开Edge浏览器失败，请检查驱动。错误信息：{}', repr(e))
            return

        success = 0
        failure = 0
        result = ''
        start = time.perf_counter()

        for i in range(0, size):
            break_flag = False

            if self.stop_status == 1:
                self.value_changed.emit(0)
                self.status.emit(1)
                self.__log('停止执行')
                break
            elif self.stop_status == 2:
                self.status.emit(2)
                self.__log('暂停执行')
                while True:
                    if self.stop_status == 0:
                        self.status.emit(0)
                        self.__log('继续执行')
                        break
                    elif self.stop_status == 1:
                        break_flag = True
                        break
                    elif self.stop_status == 2:
                        time.sleep(0.1)
            if break_flag:
                self.__log('停止执行')
                self.status.emit(1)
                break

            line = self.excel.lines[i]
            try:
                save_path = self.__screenshot(drive, line)
                self.__log('{} => 保存截图成功：{}', log.loop_msg(i + 1, size, start), save_path)
            except BaseException as e:
                traceback.print_exc()
                failure += 1
                result = 'failure:' + repr(e)
                self.__log('{} => 保存截图失败：{}', log.loop_msg(i + 1, size, start), result)
            else:
                success += 1
                result = 'success'
            finally:
                start = time.perf_counter()
                self.excel.active.cell(row=line.row, column=3, value=result)
                self.excel.xlsx.save(self.excel.path)
                self.value_changed.emit((i + 1) / size)

        self.__log('截图任务完成，成功{}条，失败{}条', success, failure)
        drive.quit()

    def run(self):
        try:
            self.exec()
        except BaseException as e:
            self.__log('未知异常：' + repr(e))
