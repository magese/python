import time
import traceback

import openpyxl
from selenium.common import StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from lrb.common import log, util


class Item:
    row = 0
    id = ''
    link = ''
    log = ''

    def __init__(self, _row, _id, _link):
        self.row = _row
        self.id = _id
        self.link = _link

    def to_string(self):
        return 'row=%s, id=%s' % (self.row, self.id)


# 读取文件内容
def read_excel(sheet):
    max_row = sheet.max_row
    max_column = sheet.max_column
    log.info('读取excel最大行数：{}，，最大列数：{}', max_row, max_column)

    items = []
    for j in range(1, max_row + 1):
        if sheet.cell(row=j, column=3).value == 'success':
            continue
        item = Item(j, sheet.cell(row=j, column=1).value, sheet.cell(row=j, column=2).value)
        items.append(item)
    return items


# 换链接
def change_link(info, edge):
    manage = util.wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="manage-list"), edge)
    id_input = util.wait_for_find_ele(
        lambda d: manage.find_element(by=By.TAG_NAME, value="input"), edge)
    id_input.send_keys('')
    id_input.clear()
    id_input.send_keys(info.id)
    id_input.send_keys(Keys.ENTER)
    time.sleep(1)

    edit_retry = 3
    while edit_retry > 0:
        try:
            edit_div = util.wait_for_find_ele(
                lambda d: d.find_element(by=By.CLASS_NAME, value="css-ece9u5"), edge)
            edit_a = util.wait_for_find_ele(
                lambda d: edit_div.find_elements(by=By.TAG_NAME, value="a"), edge)
            edit_a[0].click()
            break
        except StaleElementReferenceException:
            edit_retry -= 1
            time.sleep(0.5)

    edge.switch_to.window(edge.window_handles[-1])

    table = util.wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="el-table__fixed-body-wrapper"), edge)
    td = util.wait_for_find_ele(
        lambda d: table.find_elements(by=By.TAG_NAME, value="td"), edge)
    change_btn = util.wait_for_find_ele(
        lambda d: td[13].find_element(by=By.CLASS_NAME, value="link-text"), edge)
    change_btn.click()

    clear_btn = util.wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="css-1jjt3ne"), edge)
    clear_btn.click()
    link_input = util.wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="css-968ze5"), edge)
    link_input.clear()
    link_input.send_keys(info.link)

    save_btn = util.wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="css-php29w"), edge)
    save_btn.click()
    time.sleep(0.7)

    finish_btn = util.wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="css-r7neow"), edge)
    finish_btn.click()
    time.sleep(1)
    edge.close()
    edge.switch_to.window(edge.window_handles[0])

# main
driver = util.creative_page(False)

filepath = 'C:\\Users\\mages\\Desktop\\创意id+监测链接.xlsx'
xlsx = openpyxl.load_workbook(filepath)
active = xlsx.active
lines = read_excel(active)
size = len(lines)
log.info('读取文件成功，共读取数据{}条', size)
result = ''
T1 = time.perf_counter()

for i in range(0, size):
    line = lines[i]
    try:
        change_link(line, driver)
        log.info('{} => 换监测链接成功：{}', log.loop_msg(i + 1, size, T1), line.to_string())

    except BaseException as e:
        traceback.print_exc()
        result = 'failure:' + repr(e)
        log.info('{} => 换监测链接异常：{} => {}', log.loop_msg(i + 1, size, T1), result, line.to_string())

    else:
        result = 'success'
        time.sleep(0.5)
    finally:
        active.cell(row=line.row, column=3, value='success')
        xlsx.save(filepath)

log.info('finish')
