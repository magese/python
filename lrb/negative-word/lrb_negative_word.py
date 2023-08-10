import time
import traceback

import openpyxl
from PyQt6.QtCore import QThread, pyqtSignal
from selenium.common import StaleElementReferenceException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.webdriver import WebDriver

from lrb.common import log, util
from lrb.common.Model import Excel


class Item:
    row: int = 0
    id: str = ''
    exact_word: str = ''
    fuzzy_word: str = ''
    exact_len: int = 0
    fuzzy_len: int = 0

    def __init__(self, _row, _id, _exact_word, _fuzzy_word):
        self.row = _row
        self.id = _id
        self.exact_word = _exact_word
        self.fuzzy_word = _fuzzy_word
        self.exact_len = len(_exact_word.split('\n')) if _exact_word else 0
        self.fuzzy_len = len(_fuzzy_word.split('\n')) if _fuzzy_word else 0

    def to_string(self):
        return ('row=%s, id=%s, exact_word.size=%s, fuzzy_word.size=%s' %
                (self.row, self.id, self.exact_len, self.fuzzy_len))


class LrbNegativeWord(QThread):
    __excel_path: str = ''
    __username: str = ''
    __password: str = ''
    excel: Excel = None
    item: Item = None
    edge: WebDriver = None
    msg = pyqtSignal(str)

    def __init__(self, excel_path, username, password):
        super().__init__()
        self.__excel_path = excel_path
        self.__username = username
        self.__password = password

    # noinspection PyUnresolvedReferences
    def __emit(self, k, *p):
        msg = log.msg(k, *p)
        self.msg.emit(msg)

    def __read_excel(self):
        xlsx = openpyxl.load_workbook(self.__excel_path)
        active = xlsx.active
        max_row = active.max_row
        max_column = active.max_column
        self.__emit('读取excel最大行数：{}，最大列数：{}', max_row, max_column)

        items = []
        for j in range(1, max_row + 1):
            if str(active.cell(row=j, column=4).value).startswith('success'):
                continue
            item = Item(j, active.cell(row=j, column=1).value,
                        active.cell(row=j, column=2).value, active.cell(row=j, column=3).value)
            items.append(item)
        self.__emit('共读取待添加否定词记录 {}条', len(items))
        self.excel = Excel(self.__excel_path, xlsx, active, max_row, max_column, items)

    def __add_word(self):
        retry = 3
        is_add = False
        while retry > 0:
            try:
                manage = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="manage-list"), self.edge)
                id_input = util.wait_for_find_ele(
                    lambda d: manage.find_element(by=By.TAG_NAME, value="input"), self.edge)
                id_input.send_keys('')
                id_input.clear()
                id_input.send_keys(self.item.id)
                id_input.send_keys(Keys.ENTER)
                time.sleep(1.2)

                tbody = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="d-table__body"), self.edge)
                tr = util.wait_for_find_ele(
                    lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)
                td = util.wait_for_find_ele(
                    lambda d: tr[0].find_elements(by=By.TAG_NAME, value="td"), self.edge)
                a = util.wait_for_find_ele(
                    lambda d: td[2].find_element(by=By.TAG_NAME, value="a"), self.edge)
                a.click()
                time.sleep(1.2)

                tabs_headers = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="d-tabs-headers-linear"), self.edge)
                tabs = util.wait_for_find_ele(
                    lambda d: tabs_headers.find_elements(by=By.CLASS_NAME, value="d-tabs-header-linear"), self.edge)
                tabs[2].click()
                time.sleep(1)

                self.__de_duplicate()
                if self.item.exact_len <= 0 and self.item.fuzzy_len <= 0:
                    is_add = True
                    break

                manage = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="manage-list"), self.edge)
                btn = util.wait_for_find_ele(
                    lambda d: manage.find_elements(by=By.TAG_NAME, value="button"), self.edge)
                btn[0].click()
                time.sleep(0.5)

                radios = util.wait_for_find_ele(
                    lambda d: d.find_elements(by=By.CLASS_NAME, value="css-1b8lkqj"), self.edge)
                textarea = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="css-fx13em")
                    .find_element(by=By.TAG_NAME, value="textarea"), self.edge)
                # cancel_btn = util.wait_for_find_ele(
                #     lambda d: d.find_element(by=By.CLASS_NAME, value="css-fm44j"), self.edge)
                save_btn = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="css-r7neow"), self.edge)

                if self.item.exact_len > 0:
                    radios[0].click()
                    textarea.clear()
                    textarea.send_keys(self.item.exact_word)
                    time.sleep(1)
                    save_btn.click()

                if self.item.fuzzy_len > 0:
                    radios[1].click()
                    textarea.clear()
                    textarea.send_keys(self.item.fuzzy_word)
                    time.sleep(1)
                    save_btn.click()

                time.sleep(1)
                util.switch_unit_id_page(self.edge)
                break
            except StaleElementReferenceException:
                retry -= 1
                time.sleep(0.5)
        return is_add

    def __de_duplicate(self):
        tbody = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="d-table__body"), self.edge)
        tr = util.wait_for_find_ele(
            lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)
        if len(tr) < 2:
            return

        page_select = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="d-select-main-indicator"), self.edge)
        page_select.click()
        time.sleep(0.5)

        dropdown = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="d-dropdown-content"), self.edge)
        grid_item = util.wait_for_find_ele(
            lambda d: dropdown.find_elements(by=By.CLASS_NAME, value="d-grid-item"), self.edge)
        grid_item[len(grid_item) - 1].click()
        time.sleep(1)
        tr = util.wait_for_find_ele(
            lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)

        exact_words = self.item.exact_word.split('\n') if self.item.exact_word else []
        fuzzy_words = self.item.fuzzy_word.split('\n') if self.item.fuzzy_word else []
        for i in range(1, len(tr)):
            td = util.wait_for_find_ele(
                lambda d: tr[i].find_elements(by=By.TAG_NAME, value="td"), self.edge)
            word = td[1].get_attribute('innerText')
            meth = td[2].get_attribute('innerText')
            if meth == '精确' and word in exact_words:
                exact_words.remove(word)
            elif meth == '短语' and word in fuzzy_words:
                fuzzy_words.remove(word)
        self.item.exact_word = '\n'.join(exact_words) if len(exact_words) > 0 else None
        self.item.fuzzy_word = '\n'.join(fuzzy_words) if len(fuzzy_words) > 0 else None
        self.item.exact_len = len(exact_words)
        self.item.fuzzy_len = len(fuzzy_words)

    def exec(self):
        self.__read_excel()
        self.edge = util.unit_page(False, self.__username, self.__password)

        size = len(self.excel.lines)
        success = 0
        failure = 0
        result = ''
        start = time.perf_counter()

        for i in range(0, size):
            self.item = self.excel.lines[i]
            try:
                is_added = self.__add_word()
                note = '否定词已添加' if is_added else '否定词添加成功'
                self.__emit('{} => {}：{}', log.loop_msg(i + 1, size, start), note, self.item.to_string())

            except BaseException as e:
                traceback.print_exc()
                failure += 1
                result = 'failure:' + repr(e)
                self.__emit('{} => 添加否定词异常：{} => {}', log.loop_msg(i + 1, size, start), result, self.item.to_string())

                self.edge.quit()
                self.edge = util.unit_page(True, self.__username, self.__password)
            else:
                success += 1
                result = 'success-added' if is_added else 'success'
                time.sleep(0.5)
            finally:
                start = time.perf_counter()
                self.excel.active.cell(row=self.item.row, column=4, value=result)
                self.excel.xlsx.save(self.excel.path)

        self.__emit('添加否定词任务完成，成功{}条，失败{}条', success, failure)
        self.edge.quit()


def main():
    username = ''
    password = ''
    filepath = 'C:\\Users\\Magese\\Desktop\\批量否词需求.xlsx'
    lnw = LrbNegativeWord(filepath, username, password)
    # noinspection PyUnresolvedReferences
    lnw.msg.connect(lambda m: print(m))
    lnw.exec()


if __name__ == '__main__':
    main()
