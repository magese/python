import time

import openpyxl
from selenium.common import StaleElementReferenceException
from selenium.webdriver.common.by import By

from lrb.common import util
from lrb.common.Model import Excel, Lrb


class Item:
    row: int = 0
    id: str = ''
    word: str = ''

    def __init__(self, _row, _id, _word):
        self.row = _row
        self.id = _id
        self.word = _word

    def to_string(self):
        return 'row=%s, id=%s, word=%s' % (self.row, self.id, self.word)


class LrbSearchWord(Lrb):

    def __read_excel(self):
        xlsx = openpyxl.load_workbook(self._excel_path)
        active = xlsx.active
        max_row = active.max_row
        max_column = active.max_column
        self._emit('读取excel最大行数：{}，最大列数：{}', max_row, max_column)

        items = []
        for j in range(1, max_row + 1):
            if str(active.cell(row=j, column=4).value).startswith('success'):
                continue
            item = Item(j, active.cell(row=j, column=1).value, active.cell(row=j, column=2).value)
            items.append(item)
        self._emit('共读取待修改搜索词记录 {}条', len(items))
        self.excel = Excel(self._excel_path, xlsx, active, max_row, max_column, items)

    def __modify_word(self):
        retry = 3
        is_changed = False
        while retry > 0:
            try:
                util.search_id(self.item.id, self.edge)
                time.sleep(1.2)

                util.click_edit(self.edge)
                time.sleep(1.2)

                self.edge.switch_to.window(self.edge.window_handles[-1])

                word_input = util.wait_for_find_ele(
                    lambda d: d
                    .find_element(by=By.ID, value="ff22")
                    .find_element(by=By.TAG_NAME, value="input"), self.edge)
                word_input.send_keys('')
                word_input.clear()
                word_input.send_keys(self.item.word)
                time.sleep(1)

                finish_btn = util.wait_for_find_ele(
                    lambda d: d.find_element(by=By.CLASS_NAME, value="css-r7neow"), self.edge)
                finish_btn.click()
                time.sleep(1)
                self.edge.close()
                self.edge.switch_to.window(self.edge.window_handles[0])

                break
            except StaleElementReferenceException:
                retry -= 1
                time.sleep(0.5)
        return is_changed

    def __de_duplicate(self):
        tbody = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="d-table__body"), self.edge)
        tr = util.wait_for_find_ele(
            lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)
        if len(tr) < 2:
            return

        page_select = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="d-select-main-indicator"), self.edge)
        page_select.click()
        time.sleep(0.5)

        dropdown = util.wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="d-dropdown-content"), self.edge)
        grid_item = util.wait_for_find_ele(
            lambda d: dropdown.find_elements(by=By.CLASS_NAME, value="d-grid-item"), self.edge)
        grid_item[len(grid_item) - 1].click()
        time.sleep(1)
        tr = util.wait_for_find_ele(
            lambda d: tbody.find_elements(by=By.TAG_NAME, value="tr"), self.edge)

        exact_words = self.item.exact_word.split('\n') if self.item.exact_word else []
        fuzzy_words = self.item.fuzzy_word.split('\n') if self.item.fuzzy_word else []
        for i in range(1, len(tr)):
            td = util.wait_for_find_ele(
                lambda d: tr[i].find_elements(by=By.TAG_NAME, value="td"), self.edge)
            word = td[1].get_attribute('innerText')
            meth = td[2].get_attribute('innerText')
            if meth == '精确' and word in exact_words:
                exact_words.remove(word)
            elif meth == '短语' and word in fuzzy_words:
                fuzzy_words.remove(word)
        self.item.exact_word = '\n'.join(exact_words) if len(exact_words) > 0 else None
        self.item.fuzzy_word = '\n'.join(fuzzy_words) if len(fuzzy_words) > 0 else None
        self.item.exact_len = len(exact_words)
        self.item.fuzzy_len = len(fuzzy_words)

    def begin(self):
        super().execute(
            '修改搜索词',
            self.__read_excel,
            util.creative_page,
            self.__modify_word,
            4
        )


def main():
    filepath = 'C:\\Users\\Magese\\Desktop\\批量否词需求.xlsx'
    username = ''
    password = ''
    lsw = LrbSearchWord(filepath, username, password)
    # noinspection PyUnresolvedReferences
    lsw.msg.connect(lambda m: print(m))
    lsw.begin()


if __name__ == '__main__':
    main()
