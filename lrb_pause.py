from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.keys import Keys
import time
import openpyxl


class Item:
    row = 0
    id = ''
    log = ''

    def __init__(self, _row, _id):
        self.row = _row
        self.id = _id

    def to_string(self):
        return 'row=%s, id=%s' % (self.row, self.id)


# 等待查找元素
def wait_for_find_ele(func, edge):
    time.sleep(0.3)
    ele = WebDriverWait(edge, timeout=10).until(func)
    time.sleep(0.5)
    return ele


# 打开浏览器
def open_browser():
    options = Options()
    options.add_argument(r'--user-data-dir=C:\Users\mages\AppData\Local\Microsoft\Edge\User Data')
    # options.add_experimental_option('detach', True)
    return webdriver.Edge(options=options)


# 登录
def login(username, password, edge):
    try:
        WebDriverWait(driver, timeout=2).until(lambda d: d.find_element(by=By.CLASS_NAME, value="action-icon-group"))
    except TimeoutException:
        login_tab = wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-404bxh"), edge)
        login_tab.click()

        email_input = wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-xno39g"), edge)
        email_input.send_keys(username)

        pwd_input = wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-cct1ew"), edge)
        pwd_input.send_keys(password)

        login_btn = wait_for_find_ele(
            lambda d: d.find_element(by=By.CLASS_NAME, value="css-wp7z9d"), edge)
        login_btn.click()
    else:
        print('当前已登录')


# 切换页面
def switch_page(edge):
    menu = wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="d-new-menu__inner"), edge)
    menu_items = wait_for_find_ele(
        lambda d: menu.find_elements(by=By.CLASS_NAME, value="d-menu-item"), edge)
    menu_items[1].click()

    tabs_header = wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="d-tabs-headers-wrapper"), edge)
    tabs = wait_for_find_ele(
        lambda d: tabs_header.find_elements(by=By.CLASS_NAME, value="d-tabs-header"), edge)
    tabs[2].click()

    select = wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="d-select-wrapper"), edge)
    select.click()

    options_wrapper = wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="d-options-wrapper"), edge)
    options = wait_for_find_ele(
        lambda d: options_wrapper.find_elements(by=By.CLASS_NAME, value="d-option-name"), edge)
    options[1].click()


# 读取文件内容
def read_excel(sheet):
    max_row = sheet.max_row
    max_column = sheet.max_column
    print('最大行数：', max_row, '，最大列数：', max_column)

    items = []
    for j in range(1, max_row + 1):
        if sheet.cell(row=j, column=2).value == 'success':
            continue
        item = Item(j, sheet.cell(row=j, column=1).value)
        items.append(item)
    return items


# 暂停笔记
def pause_note(info, edge):
    manage = wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="manage-list"), edge)
    id_input = wait_for_find_ele(
        lambda d: manage.find_element(by=By.TAG_NAME, value="input"), edge)
    id_input.send_keys('')
    id_input.clear()
    id_input.send_keys(info.id)
    id_input.send_keys(Keys.ENTER)
    time.sleep(0.6)

    pause_checkbox = wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="css-1a4w089"), edge)
    checked = pause_checkbox.get_attribute('data-is-checked')

    pause_switch = wait_for_find_ele(
        lambda d: d.find_element(by=By.CLASS_NAME, value="css-vwjppn"), edge)

    if 'true' == checked:
        pause_switch.click()
    else:
        print('笔记：', info.id, '已暂停')
    time.sleep(0.5)


# main
driver = open_browser()
print('打开Edge成功')
driver.get("https://ad.xiaohongshu.com/")
time.sleep(2)

login('print(current_time_str(),', '', driver)
print('登录账号成功')
time.sleep(2)

switch_page(driver)
print('切换页面成功')
time.sleep(2)

filepath = 'C:\\Users\\mages\\Desktop\\暂停创意id.xlsx'
xlsx = openpyxl.load_workbook(filepath)
active = xlsx.active
lines = read_excel(active)
size = len(lines)
print('读取文件成功，共读取数据', size, "条")

for i in range(0, size):
    line = lines[i]
    try:
        pause_note(line, driver)
        print(i + 1, '/', size, '-', format((i + 1) / size * 100, '.2f'), '% => ', line.to_string())
        time.sleep(0.5)
    except BaseException as e:
        value = 'failure:' + str(e)
        print('暂停笔记异常：', value, line.to_string())
        active.cell(row=line.row, column=2, value=value)
        driver.quit()
        driver = open_browser()
        print('重新打开Edge成功')
        driver.get("https://ad.xiaohongshu.com/")
        time.sleep(2)
        login('print(current_time_str(),', '', driver)
        print('重新登录账号成功')
        time.sleep(2)
        switch_page(driver)
        print('重新切换页面成功')
        time.sleep(2)
    else:
        active.cell(row=line.row, column=2, value='success')
    finally:
        xlsx.save(filepath)
print('finish')
